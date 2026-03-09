#!/usr/bin/env python3
"""
Fronius Excel → MSSQL Price Sync (Faz 1)

Reads Fronius price list Excel, matches products against MSSQL URUNLER table,
and updates FIYAT1 (sell price) and ALISFIYATI (buy price).

Faz 1: ONLY price updates for existing matched products. No INSERT/DELETE/STOCK changes.

Usage:
  python3 fronius_sync.py                      # dry-run (default)
  python3 fronius_sync.py --apply              # actually write to DB
  python3 fronius_sync.py --excel path.xlsx    # custom Excel path
"""

import argparse
import csv
import logging
import os
import re
import sys
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
import pymssql
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent
EXCEL_DEFAULT = SCRIPT_DIR.parent / "Fronius_Fiyat_Listesi_FSP_2026.xlsx"
REPORT_CSV = SCRIPT_DIR / "fronius_sync_report.csv"
MAIN_LOG = SCRIPT_DIR / "sync_log.txt"
LOG_DIR = SCRIPT_DIR / "logs"

MARKUP = 1.20  # FIYAT1 = excel_price * MARKUP
FRONIUS_MARKA_ID = 49
FUZZY_THRESHOLD = 0.70  # min similarity for fuzzy match candidate

# Column groups in Excel: (name_col, sku_col, price_col) — 0-indexed
COLUMN_GROUPS = [
    (0, 1, 2),    # A, B, C
    (5, 6, 7),    # F, G, H
    (10, 11, 12), # K, L, M
]

# Rows that are headers / section labels (detected dynamically)
HEADER_KEYWORDS = {"ürün adı", "urun adi", "fronius fiyat listesi"}

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
LOG_DIR.mkdir(exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
detail_log_path = LOG_DIR / f"fronius_sync_{timestamp}.log"

logger = logging.getLogger("fronius_sync")
logger.setLevel(logging.DEBUG)

# File handler — detailed
fh = logging.FileHandler(detail_log_path, encoding="utf-8")
fh.setLevel(logging.DEBUG)
fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
logger.addHandler(fh)

# Console handler — info+
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
ch.setFormatter(logging.Formatter("%(message)s"))
logger.addHandler(ch)

# Append to main sync log
mlh = logging.FileHandler(MAIN_LOG, encoding="utf-8")
mlh.setLevel(logging.INFO)
mlh.setFormatter(logging.Formatter("%(asctime)s %(message)s"))
logger.addHandler(mlh)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def normalize_sku(raw: str) -> str:
    """Normalize SKU: remove commas, spaces, dashes, dots. Uppercase."""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[,\s\.\-]", "", s)
    return s.upper()


def normalize_turkish(s: str) -> str:
    """Replace Turkish special chars with ASCII equivalents."""
    tr_map = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU")
    return s.translate(tr_map)


def normalize_name(raw: str) -> str:
    """Normalize product name: lowercase, Turkish→ASCII, strip extra spaces/punctuation."""
    if not raw:
        return ""
    s = str(raw).strip().lower()
    s = normalize_turkish(s)
    # Remove common decorative chars
    s = s.replace("–", " ").replace("-", " ").replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_header_row(name_val) -> bool:
    """Check if a cell value is a header/label, not a product."""
    if not name_val:
        return True
    n = str(name_val).strip().lower()
    if n in HEADER_KEYWORDS:
        return True
    # Section headers don't have a numeric price next to them
    return False


def fuzzy_score(a: str, b: str) -> float:
    """Return similarity ratio between two normalized strings."""
    return SequenceMatcher(None, a, b).ratio()


# ---------------------------------------------------------------------------
# Excel Parser
# ---------------------------------------------------------------------------
def parse_excel(path: str) -> list[dict]:
    """
    Parse Fronius price list Excel.
    Returns list of {name, sku, sku_norm, name_norm, price}.
    """
    logger.info(f"Reading Excel: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    logger.info(f"  Sheet: {ws.title}, rows={ws.max_row}, cols={ws.max_column}")

    products = []
    seen_skus = set()

    for row_idx in range(1, ws.max_row + 1):
        row_cells = [ws.cell(row=row_idx, column=c + 1).value for c in range(ws.max_column)]

        for name_col, sku_col, price_col in COLUMN_GROUPS:
            name_val = row_cells[name_col] if name_col < len(row_cells) else None
            sku_val = row_cells[sku_col] if sku_col < len(row_cells) else None
            price_val = row_cells[price_col] if price_col < len(row_cells) else None

            # Skip empty / header rows
            if not name_val or not sku_val or price_val is None:
                continue
            if is_header_row(name_val):
                continue

            # Price must be numeric
            try:
                price = float(price_val)
            except (ValueError, TypeError):
                logger.debug(f"  Row {row_idx}: non-numeric price {repr(price_val)} for {name_val}")
                continue

            if price <= 0:
                continue

            sku_norm = normalize_sku(str(sku_val))
            if not sku_norm:
                continue

            # Deduplicate by normalized SKU
            if sku_norm in seen_skus:
                logger.debug(f"  Duplicate SKU {sku_norm} at row {row_idx}, skipping")
                continue
            seen_skus.add(sku_norm)

            products.append({
                "name": str(name_val).strip(),
                "sku": str(sku_val).strip(),
                "sku_norm": sku_norm,
                "name_norm": normalize_name(name_val),
                "price": round(price, 2),
            })

    logger.info(f"  Parsed {len(products)} products from Excel")
    return products


# ---------------------------------------------------------------------------
# DB Functions
# ---------------------------------------------------------------------------
def get_db_connection():
    """Create MSSQL connection from .env credentials."""
    load_dotenv(SCRIPT_DIR / ".env")

    server = os.getenv("MSSQL_SERVER")
    port = int(os.getenv("MSSQL_PORT", "1433"))
    user = os.getenv("MSSQL_USER")
    password = os.getenv("MSSQL_PASSWORD")
    database = os.getenv("MSSQL_DATABASE")

    if not all([server, user, password, database]):
        logger.error("Missing DB credentials in .env")
        sys.exit(1)

    logger.info(f"Connecting to MSSQL {server}:{port}/{database}")
    conn = pymssql.connect(
        server=server,
        port=port,
        user=user,
        password=password,
        database=database,
        charset="utf8",
    )
    logger.info("  Connected OK")
    return conn


def fetch_fronius_products(conn) -> list[dict]:
    """Fetch all Fronius products from URUNLER (MARKAID=49)."""
    cursor = conn.cursor(as_dict=True)
    cursor.execute("""
        SELECT
            ID, URUNADI, STOKKODU, STOK, FIYAT1, ALISFIYATI
        FROM URUNLER
        WHERE MARKAID = %s
    """, (FRONIUS_MARKA_ID,))
    rows = cursor.fetchall()
    cursor.close()

    # Enrich with normalized fields
    for r in rows:
        r["stokkodu_norm"] = normalize_sku(r["STOKKODU"] or "")
        r["urunadi_norm"] = normalize_name(r["URUNADI"] or "")

    logger.info(f"  Fetched {len(rows)} Fronius products from DB")
    return rows


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------
def match_products(excel_products: list[dict], db_products: list[dict]) -> list[dict]:
    """
    Match Excel products to DB products.
    Returns list of result dicts for the CSV report.
    """
    results = []

    # Build lookup indexes
    db_by_sku = {}
    for p in db_products:
        if p["stokkodu_norm"]:
            db_by_sku[p["stokkodu_norm"]] = p

    db_by_name = {}
    for p in db_products:
        if p["urunadi_norm"]:
            db_by_name[p["urunadi_norm"]] = p

    matched_db_ids = set()

    for ep in excel_products:
        result = {
            "excel_sku": ep["sku"],
            "excel_name": ep["name"],
            "excel_price": ep["price"],
            "db_product_id": "",
            "db_stokkodu": "",
            "db_name": "",
            "db_stock": "",
            "old_fiyat1": "",
            "new_fiyat1": "",
            "old_alisfiyati": "",
            "new_alisfiyati": "",
            "action": "",
            "reason": "",
        }

        # 1. Try exact SKU match
        db_match = db_by_sku.get(ep["sku_norm"])
        match_method = "sku_exact" if db_match else None

        # 2. Try exact name contains match
        if not db_match:
            for dp in db_products:
                if dp["ID"] in matched_db_ids:
                    continue
                if ep["name_norm"] and dp["urunadi_norm"]:
                    # Check if excel name is contained in DB name or vice versa
                    if ep["name_norm"] in dp["urunadi_norm"] or dp["urunadi_norm"] in ep["name_norm"]:
                        db_match = dp
                        match_method = "name_exact"
                        break

        # 3. Fuzzy match — only for reporting
        fuzzy_candidate = None
        if not db_match:
            best_score = 0
            for dp in db_products:
                if dp["ID"] in matched_db_ids:
                    continue
                score = fuzzy_score(ep["name_norm"], dp["urunadi_norm"])
                if score > best_score and score >= FUZZY_THRESHOLD:
                    best_score = score
                    fuzzy_candidate = (dp, score)

        if db_match:
            matched_db_ids.add(db_match["ID"])
            stock = db_match["STOK"] or 0
            old_fiyat1 = db_match["FIYAT1"] or 0
            old_alis = db_match["ALISFIYATI"] or 0

            result["db_product_id"] = db_match["ID"]
            result["db_stokkodu"] = db_match["STOKKODU"] or ""
            result["db_name"] = db_match["URUNADI"] or ""
            result["db_stock"] = stock
            result["old_fiyat1"] = round(float(old_fiyat1), 2)
            result["old_alisfiyati"] = round(float(old_alis), 2)

            new_alis = round(ep["price"], 2)
            result["new_alisfiyati"] = new_alis

            if stock > 0:
                new_fiyat1 = round(ep["price"] * MARKUP, 2)
                result["new_fiyat1"] = new_fiyat1
                result["action"] = "UPDATE_PRICE"
                result["reason"] = f"match={match_method}, stock={stock}"
            else:
                result["new_fiyat1"] = 0
                result["action"] = "ZERO_PRICE_STOCK0"
                result["reason"] = f"match={match_method}, stock=0 → FIYAT1=0"
        elif fuzzy_candidate:
            dp, score = fuzzy_candidate
            result["db_product_id"] = dp["ID"]
            result["db_stokkodu"] = dp["STOKKODU"] or ""
            result["db_name"] = dp["URUNADI"] or ""
            result["db_stock"] = dp["STOK"] or 0
            result["old_fiyat1"] = round(float(dp["FIYAT1"] or 0), 2)
            result["old_alisfiyati"] = round(float(dp["ALISFIYATI"] or 0), 2)
            result["action"] = "MANUAL_REVIEW"
            result["reason"] = f"fuzzy={score:.2f}, no exact match"
        else:
            result["action"] = "NOT_IN_DB"
            result["reason"] = "no match found in DB"

        results.append(result)

    # DB products not in Excel
    for dp in db_products:
        if dp["ID"] not in matched_db_ids:
            results.append({
                "excel_sku": "",
                "excel_name": "",
                "excel_price": "",
                "db_product_id": dp["ID"],
                "db_stokkodu": dp["STOKKODU"] or "",
                "db_name": dp["URUNADI"] or "",
                "db_stock": dp["STOK"] or 0,
                "old_fiyat1": round(float(dp["FIYAT1"] or 0), 2),
                "new_fiyat1": "",
                "old_alisfiyati": round(float(dp["ALISFIYATI"] or 0), 2),
                "new_alisfiyati": "",
                "action": "SKIP",
                "reason": "DB product not in Excel",
            })

    return results


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------
CSV_COLUMNS = [
    "excel_sku", "excel_name", "excel_price",
    "db_product_id", "db_stokkodu", "db_name", "db_stock",
    "old_fiyat1", "new_fiyat1", "old_alisfiyati", "new_alisfiyati",
    "action", "reason",
]


def write_report(results: list[dict], path: Path):
    """Write CSV report."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(results)
    logger.info(f"  Report written: {path}")


def print_summary(results: list[dict]):
    """Print action summary to console."""
    from collections import Counter
    counts = Counter(r["action"] for r in results)

    logger.info("")
    logger.info("=" * 60)
    logger.info("SYNC SUMMARY")
    logger.info("=" * 60)
    for action in ["UPDATE_PRICE", "ZERO_PRICE_STOCK0", "MANUAL_REVIEW", "NOT_IN_DB", "SKIP"]:
        if counts.get(action, 0) > 0:
            logger.info(f"  {action:25s}: {counts[action]}")
    logger.info(f"  {'TOTAL':25s}: {len(results)}")
    logger.info("=" * 60)

    # Show price changes detail
    updates = [r for r in results if r["action"] in ("UPDATE_PRICE", "ZERO_PRICE_STOCK0")]
    if updates:
        logger.info("")
        logger.info("PRICE CHANGES:")
        for r in updates:
            old_f = r["old_fiyat1"]
            new_f = r["new_fiyat1"]
            change = ""
            if old_f and new_f and old_f != new_f:
                diff = float(new_f) - float(old_f)
                change = f" (diff: {diff:+.2f})"
            logger.info(
                f"  [{r['action']}] {r['db_name'][:50]:50s} "
                f"FIYAT1: {old_f} → {new_f}{change}"
            )

    # Show manual review candidates
    reviews = [r for r in results if r["action"] == "MANUAL_REVIEW"]
    if reviews:
        logger.info("")
        logger.info("MANUAL REVIEW CANDIDATES:")
        for r in reviews:
            logger.info(
                f"  Excel: {r['excel_name'][:40]:40s} ({r['excel_sku']})"
            )
            logger.info(
                f"  DB:    {r['db_name'][:40]:40s} ({r['db_stokkodu']})"
            )
            logger.info(f"  Reason: {r['reason']}")
            logger.info("")


# ---------------------------------------------------------------------------
# Apply
# ---------------------------------------------------------------------------
def apply_updates(conn, results: list[dict]):
    """Apply price updates to DB in a single transaction."""
    updates = [r for r in results if r["action"] in ("UPDATE_PRICE", "ZERO_PRICE_STOCK0")]

    if not updates:
        logger.info("No updates to apply.")
        return

    logger.info(f"Applying {len(updates)} price updates...")

    cursor = conn.cursor()
    try:
        for r in updates:
            product_id = r["db_product_id"]
            new_fiyat1 = r["new_fiyat1"]
            new_alis = r["new_alisfiyati"]

            cursor.execute("""
                UPDATE URUNLER
                SET FIYAT1 = %s, ALISFIYATI = %s
                WHERE ID = %s
            """, (new_fiyat1, new_alis, product_id))

            logger.debug(
                f"  Updated ID={product_id}: FIYAT1={new_fiyat1}, ALISFIYATI={new_alis}"
            )

        conn.commit()
        logger.info(f"  COMMITTED {len(updates)} updates successfully")

    except Exception as e:
        conn.rollback()
        logger.error(f"  ROLLBACK — error during apply: {e}")
        raise


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Fronius Excel → MSSQL Price Sync")
    parser.add_argument("--apply", action="store_true", help="Actually write to DB (default: dry-run)")
    parser.add_argument("--excel", type=str, default=str(EXCEL_DEFAULT), help="Path to Excel file")
    args = parser.parse_args()

    mode = "APPLY" if args.apply else "DRY-RUN"
    logger.info(f"=== Fronius Price Sync — {mode} — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===")
    logger.info(f"  Markup: {MARKUP}x")
    logger.info(f"  Excel: {args.excel}")
    logger.info(f"  Detail log: {detail_log_path}")

    # 1. Parse Excel
    excel_products = parse_excel(args.excel)
    if not excel_products:
        logger.error("No products found in Excel. Aborting.")
        sys.exit(1)

    # 2. Connect to DB
    conn = get_db_connection()

    try:
        # 3. Fetch DB products
        db_products = fetch_fronius_products(conn)

        # 4. Match
        results = match_products(excel_products, db_products)

        # 5. Report
        write_report(results, REPORT_CSV)
        print_summary(results)

        # 6. Apply if requested
        if args.apply:
            logger.info("")
            logger.info("*** APPLY MODE — writing to database ***")
            apply_updates(conn, results)
        else:
            logger.info("")
            logger.info("DRY-RUN mode — no changes written to DB.")
            logger.info(f"Review report: {REPORT_CSV}")
            logger.info("Run with --apply to execute updates.")

    finally:
        conn.close()
        logger.info("DB connection closed.")

    logger.info(f"Done. Log: {detail_log_path}")


if __name__ == "__main__":
    main()
