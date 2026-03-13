#!/usr/bin/env python3
"""
Fronius Stock Tracker
- Downloads stock Excel from Fronius Filebox (Nextcloud)
- Compares with previous stock data
- Sends HTML email via SendGrid with changes

Usage:
  python3 fronius_stock_tracker.py
"""

import json
import os
import re
import sys
import tempfile
import zipfile
from datetime import datetime
from glob import glob
from pathlib import Path
from urllib.request import Request, urlopen

import openpyxl

SCRIPT_DIR = Path(__file__).parent
DATA_FILE = SCRIPT_DIR / "previous_stock.json"
FILEBOX_SHARE_TOKEN = "jQPg72Ljk43MsmE"
FILEBOX_PASSWORD = "Fronius2026."
FILEBOX_BASE = "https://filebox.fronius.com"

SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
if not SENDGRID_API_KEY:
    print("ERROR: SENDGRID_API_KEY environment variable is required")
    sys.exit(1)
EMAIL_FROM = "Turkiye Solar Market <support@keywork.ai>"
EMAIL_TO = "destek@turkiyesolarmarket.com.tr"

# Name aliases: map old/variant names → canonical name
# When Fronius renames a product in their Excel, add the mapping here
# so the tracker treats it as the same product (not disappear+appear)
NAME_ALIASES = {
    # Fronius price list says "33.3", stock Excel says "33.0"
    # If stock Excel switches to "33.3", map back to canonical "33.0"
    "Verto 33.3 SPD 1+2": "Verto 33.0 SPD 1+2",
    "Verto 33.3 Plus SPD 1+2": "Verto 33.0 Plus SPD 1+2",
}


def download_stock_excel():
    """Download stock Excel from Fronius Filebox via Nextcloud auth."""
    import http.cookiejar
    from urllib.parse import urlencode

    cj = http.cookiejar.CookieJar()
    opener = __import__("urllib.request", fromlist=["build_opener"]).build_opener(
        __import__("urllib.request", fromlist=["HTTPCookieProcessor"]).HTTPCookieProcessor(cj)
    )

    # Step 1: Get auth page + CSRF token
    auth_url = f"{FILEBOX_BASE}/s/{FILEBOX_SHARE_TOKEN}"
    req = Request(auth_url)
    resp = opener.open(req)
    html = resp.read().decode("utf-8")

    m = re.search(r'data-requesttoken="([^"]+)"', html)
    if not m:
        raise RuntimeError("Could not find requesttoken on Filebox auth page")
    token = m.group(1)

    # Step 2: Authenticate with password
    auth_post_url = f"{FILEBOX_BASE}/s/{FILEBOX_SHARE_TOKEN}/authenticate/showshare"
    data = urlencode({
        "password": FILEBOX_PASSWORD,
        "sharingToken": FILEBOX_SHARE_TOKEN,
    }).encode()
    req = Request(auth_post_url, data=data, headers={"requesttoken": token})
    resp = opener.open(req)

    # Step 3: Download ZIP
    dl_url = f"{FILEBOX_BASE}/s/{FILEBOX_SHARE_TOKEN}/download"
    req = Request(dl_url)
    with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp:
        resp = opener.open(req)
        tmp.write(resp.read())
        zip_path = tmp.name

    # Step 4: Extract Excel from ZIP
    excel_path = None
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if name.endswith(".xlsx"):
                excel_path = zf.extract(name, tempfile.gettempdir())
                break

    os.unlink(zip_path)
    if not excel_path:
        raise RuntimeError("No .xlsx file found in Filebox ZIP")

    return excel_path


def normalize_name(name):
    """Normalize product name: collapse whitespace, apply aliases."""
    name = re.sub(r'\s+', ' ', str(name).strip())
    return NAME_ALIASES.get(name, name)


def parse_stock_excel(path):
    """Parse Fronius stock Excel. Returns dict {product_name: quantity}."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    stock = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[0]
        qty = row[1]
        if name and qty is not None:
            stock[normalize_name(name)] = int(qty)
    return stock


def load_previous_stock():
    """Load previous stock from JSON file."""
    if DATA_FILE.exists():
        data = json.loads(DATA_FILE.read_text("utf-8"))
        raw = data.get("stock", {})
        # Re-normalize old keys to match current normalization
        stock = {}
        for k, v in raw.items():
            stock[normalize_name(k)] = v
        return stock, data.get("date", "bilinmiyor")
    return None, None


def save_current_stock(stock):
    """Save current stock to JSON file."""
    data = {
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "stock": stock,
    }
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")


def build_email_html(current, previous, prev_date, excel_filename):
    """Build comparison HTML email."""
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    all_products = sorted(set(list(current.keys()) + list(previous.keys() if previous else [])))

    in_stock = []
    out_of_stock = []
    changed = []

    for name in all_products:
        cur_qty = current.get(name, 0)
        prev_qty = previous.get(name, 0) if previous else None

        if cur_qty > 0:
            in_stock.append((name, cur_qty, prev_qty))
        else:
            out_of_stock.append((name, cur_qty, prev_qty))

        if previous and cur_qty != prev_qty:
            changed.append((name, prev_qty, cur_qty))

    # Stats
    total = len(all_products)
    total_in = len(in_stock)
    total_out = len(out_of_stock)
    total_changed = len(changed)

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="font-family: Arial, sans-serif; color: #333; max-width: 800px; margin: 0 auto; padding: 20px;">

<h2 style="color: #1a5276; border-bottom: 2px solid #1a5276; padding-bottom: 10px;">
  Fronius Stok Raporu - {now}
</h2>

<p style="color: #666; font-size: 14px;">
  Kaynak: {excel_filename}<br>
  {"Onceki kontrol: " + prev_date if prev_date else "Ilk kontrol (onceki veri yok)"}
</p>

<table style="width: 100%; border-collapse: collapse; margin: 15px 0;">
  <tr>
    <td style="padding: 10px; background: #d4efdf; border-radius: 5px; text-align: center; font-size: 18px; font-weight: bold; color: #1e8449;">
      {total_in} Stokta
    </td>
    <td style="width: 10px;"></td>
    <td style="padding: 10px; background: #fadbd8; border-radius: 5px; text-align: center; font-size: 18px; font-weight: bold; color: #c0392b;">
      {total_out} Stok Yok
    </td>
    <td style="width: 10px;"></td>
    <td style="padding: 10px; background: #fef9e7; border-radius: 5px; text-align: center; font-size: 18px; font-weight: bold; color: #b7950b;">
      {total_changed} Degisiklik
    </td>
  </tr>
</table>"""

    # Changes section (only if there's previous data)
    if previous and changed:
        html += """
<h3 style="color: #b7950b; margin-top: 25px;">Degisiklikler</h3>
<table style="width: 100%; border-collapse: collapse; font-size: 14px;">
  <tr style="background: #f8f9fa;">
    <th style="padding: 8px; text-align: left; border-bottom: 2px solid #dee2e6;">Urun</th>
    <th style="padding: 8px; text-align: center; border-bottom: 2px solid #dee2e6;">Onceki</th>
    <th style="padding: 8px; text-align: center; border-bottom: 2px solid #dee2e6;">Guncel</th>
    <th style="padding: 8px; text-align: center; border-bottom: 2px solid #dee2e6;">Durum</th>
  </tr>"""
        for name, prev_qty, cur_qty in sorted(changed, key=lambda x: x[2] - x[1]):
            diff = cur_qty - prev_qty
            if cur_qty == 0 and prev_qty > 0:
                badge = '<span style="color:#c0392b; font-weight:bold;">STOK BITTI</span>'
            elif prev_qty == 0 and cur_qty > 0:
                badge = '<span style="color:#1e8449; font-weight:bold;">STOGA GIRDI</span>'
            elif diff > 0:
                badge = f'<span style="color:#1e8449;">+{diff}</span>'
            else:
                badge = f'<span style="color:#c0392b;">{diff}</span>'
            html += f"""
  <tr style="border-bottom: 1px solid #eee;">
    <td style="padding: 6px 8px;">{name}</td>
    <td style="padding: 6px 8px; text-align: center;">{prev_qty}</td>
    <td style="padding: 6px 8px; text-align: center;">{cur_qty}</td>
    <td style="padding: 6px 8px; text-align: center;">{badge}</td>
  </tr>"""
        html += "</table>"

    # In Stock
    html += """
<h3 style="color: #1e8449; margin-top: 25px;">Stokta Olan Urunler</h3>
<table style="width: 100%; border-collapse: collapse; font-size: 14px;">
  <tr style="background: #f8f9fa;">
    <th style="padding: 8px; text-align: left; border-bottom: 2px solid #dee2e6;">Urun</th>
    <th style="padding: 8px; text-align: center; border-bottom: 2px solid #dee2e6;">Adet</th>
  </tr>"""
    for name, qty, _ in in_stock:
        html += f"""
  <tr style="border-bottom: 1px solid #eee;">
    <td style="padding: 6px 8px;">{name}</td>
    <td style="padding: 6px 8px; text-align: center; font-weight: bold; color: #1e8449;">{qty}</td>
  </tr>"""
    html += "</table>"

    # Out of Stock
    html += """
<h3 style="color: #c0392b; margin-top: 25px;">Stokta Olmayan Urunler</h3>
<table style="width: 100%; border-collapse: collapse; font-size: 14px;">
  <tr style="background: #f8f9fa;">
    <th style="padding: 8px; text-align: left; border-bottom: 2px solid #dee2e6;">Urun</th>
    <th style="padding: 8px; text-align: center; border-bottom: 2px solid #dee2e6;">Adet</th>
  </tr>"""
    for name, qty, _ in out_of_stock:
        html += f"""
  <tr style="border-bottom: 1px solid #eee;">
    <td style="padding: 6px 8px;">{name}</td>
    <td style="padding: 6px 8px; text-align: center; color: #c0392b;">0</td>
  </tr>"""
    html += "</table>"

    html += f"""
<hr style="margin-top: 30px; border: none; border-top: 1px solid #eee;">
<p style="color: #999; font-size: 12px;">
  Bu rapor otomatik olusturulmustur. Toplam {total} urun takip edilmektedir.
</p>
</body></html>"""

    return html


def send_email(subject, html_body):
    """Send email via SendGrid API."""
    payload = {
        "personalizations": [{"to": [{"email": EMAIL_TO}]}],
        "from": {"email": "support@keywork.ai", "name": "Turkiye Solar Market"},
        "subject": subject,
        "content": [{"type": "text/html", "value": html_body}],
    }

    data = json.dumps(payload).encode("utf-8")
    req = Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=data,
        headers={
            "Authorization": f"Bearer {SENDGRID_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        resp = urlopen(req)
        print(f"Email sent: HTTP {resp.status}")
    except Exception as e:
        print(f"Email send failed: {e}")
        if hasattr(e, "read"):
            print(e.read().decode())
        raise


def main():
    print("=== Fronius Stock Tracker ===")
    print()

    # 1. Download Excel
    print("Downloading stock Excel from Fronius Filebox...")
    excel_path = download_stock_excel()
    excel_filename = os.path.basename(excel_path)
    print(f"  Downloaded: {excel_filename}")

    # 2. Parse
    print("Parsing stock data...")
    current_stock = parse_stock_excel(excel_path)
    print(f"  Found {len(current_stock)} products")

    # 3. Load previous
    previous_stock, prev_date = load_previous_stock()
    if previous_stock:
        print(f"  Previous stock from: {prev_date} ({len(previous_stock)} products)")
    else:
        print("  No previous stock data (first run)")

    # 4. Compare
    changes = 0
    if previous_stock:
        for name in set(list(current_stock.keys()) + list(previous_stock.keys())):
            if current_stock.get(name, 0) != previous_stock.get(name, 0):
                changes += 1
        print(f"  Changes detected: {changes}")

    # 5. Build + send email
    in_stock = sum(1 for v in current_stock.values() if v > 0)
    out_stock = sum(1 for v in current_stock.values() if v == 0)
    now = datetime.now().strftime("%d.%m.%Y")

    subject = f"Fronius Stok Raporu {now} - {in_stock} stokta, {out_stock} yok"
    if changes > 0:
        subject += f", {changes} degisiklik"

    print(f"Sending email to {EMAIL_TO}...")
    html = build_email_html(current_stock, previous_stock, prev_date, excel_filename)
    send_email(subject, html)

    # 6. Save current as previous
    save_current_stock(current_stock)
    print(f"  Stock data saved to {DATA_FILE}")

    # Cleanup
    try:
        os.unlink(excel_path)
    except OSError:
        pass

    print("\nDone!")


if __name__ == "__main__":
    main()
